import openpyxl
from openpyxl import load_workbook

def clean_cell_range(path, sheet_name: str, cell_range, target: str, replacement: str):
    replacements = 0
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f'[{sheet_name}] not found. Quitting.')
        return

    sheet = workbook[sheet_name]
    for column in sheet[cell_range]:
        for cell in column:
            if cell.value == target:
                cell.value = replacement
                replacements += 1
    
    print(f'Found {replacements} instances of "{target}" and replaced with "{replacement}"')


if __name__ == "__main__":

    target_workbook = 'cirrus8/Cirrus8 Tenancy Schedule (Compact) - January 2024.xlsx'
    clean_cell_range(path=target_workbook, sheet_name='Page 8', 
                     cell_range='O10:Q39', target='Infinity', replacement='0')